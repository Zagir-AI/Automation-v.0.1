#!/usr/bin/env python3
"""
cli.py — главная точка входа в систему.

Использование:
  python cli.py new <код_объекта> "<название>"         — создать новый проект
  python cli.py calc <путь>                            — рассчитать нагрузки ВРУ
  python cli.py calc-outdoor <путь>                   — рассчитать наружные сети
  python cli.py summary <путь>                         — краткая сводка
  python cli.py docs <путь> [--type spec|cable|load|pnr]  — документы
  python cli.py plan <путь> [--section ОВ]            — DXF-план
  python cli.py check-selectivity <путь>              — проверить селективность
  python cli.py check-compensation <путь>             — проверить КРМ
  python cli.py import <путь> dwg|table <файл>        — импорт из DXF или Excel
  python cli.py validate <путь>                        — проверить JSON
  python cli.py change <путь> [--desc "..."] [--reason 01-04] — зарегистрировать ревизию
  python cli.py list                                   — список проектов
"""

import sys
import os
import json
import argparse
import copy
from pathlib import Path

# ── пути ────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
PROJECTS_DIR = ROOT / "projects"
PROJECTS_DIR.mkdir(exist_ok=True)
sys.path.insert(0, str(ROOT))

# ── ANSI цвета для терминала ──────────────────────────────────────────
class C:
    RESET  = "\033[0m"
    BOLD   = "\033[1m"
    GREEN  = "\033[92m"
    YELLOW = "\033[93m"
    RED    = "\033[91m"
    CYAN   = "\033[96m"
    GRAY   = "\033[90m"
    WHITE  = "\033[97m"

def ok(s):   return f"{C.GREEN}✓{C.RESET} {s}"
def warn(s): return f"{C.YELLOW}⚠{C.RESET} {s}"
def err(s):  return f"{C.RED}✗{C.RESET} {s}"
def info(s): return f"{C.CYAN}→{C.RESET} {s}"
def hdr(s):  return f"\n{C.BOLD}{C.WHITE}{s}{C.RESET}"


# ── УТИЛИТЫ ──────────────────────────────────────────────────────────

def load_project(path: Path) -> dict:
    json_file = path / "project.json" if path.is_dir() else path
    if not json_file.exists():
        print(err(f"Файл не найден: {json_file}"))
        sys.exit(1)
    with open(json_file, encoding="utf-8") as f:
        return json.load(f)

def save_project(project: dict, path: Path):
    json_file = path / "project.json" if path.is_dir() else path
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(project, f, ensure_ascii=False, indent=2)

def find_project_dir(arg: str) -> Path:
    """Найти папку проекта по аргументу (путь или код)."""
    p = Path(arg)
    if p.exists():
        return p
    for d in PROJECTS_DIR.iterdir():
        if d.is_dir() and arg.upper() in d.name.upper():
            return d
    print(err(f"Проект не найден: {arg}"))
    sys.exit(1)

def _ensure_calc(project: dict, proj_dir: Path) -> dict:
    """Если расчёт не выполнен — запускает автоматически."""
    if not project.get("_results"):
        print(info("Запускаю расчёт автоматически..."))
        from calc.engine import calculate_project
        project = calculate_project(project)
        save_project(project, proj_dir)
    return project


# ── КОМАНДЫ ───────────────────────────────────────────────────────────

def cmd_list(args):
    """Список всех проектов."""
    print(hdr("Проекты"))
    projects = sorted(PROJECTS_DIR.iterdir())
    if not projects:
        print(info("Нет проектов. Создай первый: python cli.py new КОД 'Название'"))
        return

    fmt = f"  {{:<30}} {{:<12}} {{:<8}} {{}}"
    print(C.GRAY + fmt.format("Папка", "Код", "Стадия", "Название") + C.RESET)
    print(C.GRAY + "  " + "─"*70 + C.RESET)

    for d in projects:
        if not d.is_dir():
            continue
        jf = d / "project.json"
        if not jf.exists():
            continue
        try:
            p = json.loads(jf.read_text(encoding="utf-8"))["project"]
            calc = " [✓расчёт]" if (d / "project.json").read_text().__contains__('"calc_done": true') else ""
            print(fmt.format(
                d.name[:30],
                p.get("code","")[:12],
                p.get("stage",""),
                p.get("name","")[:40] + calc
            ))
        except Exception:
            print(f"  {d.name} — ошибка чтения")


def _prompt(label: str, default: str = "", required: bool = False) -> str:
    """Интерактивный ввод одного поля штампа. Enter = оставить default.
    В не-TTY (CI/скрипты/EOF) — без падения возвращает default."""
    suffix = f" [{default}]" if default else ""
    while True:
        try:
            val = input(f"  {label}{suffix}: ").strip()
        except EOFError:
            return default
        if not val:
            val = default
        if val or not required:
            return val
        print(f"  {C.RED}Обязательное поле — введи значение{C.RESET}")


def _input_stamp(current: dict | None = None) -> dict:
    """
    Интерактивный ввод штампа проекта.
    current — текущие значения (для режима редактирования).
    """
    c = current or {}
    print(f"\n{C.CYAN}Заполни штамп проекта (Enter — оставить текущее):{C.RESET}")
    return {
        "designer":   _prompt("Разработал (ФИО)",      c.get("designer", "")),
        "checker":    _prompt("Проверил (ФИО)",         c.get("checker", "")),
        "norm_head":  _prompt("Н.контроль (ФИО)",       c.get("norm_head", "")),
        "gip":        _prompt("ГИП (ФИО)",              c.get("gip", "")),
        "gap":        _prompt("ГАП (ФИО, необяз.)",     c.get("gap", "")),
        "org":        _prompt("Организация",             c.get("org", "")),
        "city":       _prompt("Город",                   c.get("city", "")),
        "stage":      _prompt("Стадия [Р]",              c.get("stage", "Р")) or "Р",
        "object_type":_prompt("Тип объекта",             c.get("object_type","Общественное здание")),
        "system":     _prompt("Система заземления [TN-S]",c.get("system","TN-S")) or "TN-S",
    }


def cmd_new(args):
    """Создать новый проект из шаблона.

    Режимы:
      - Интерактивный (TTY и без флагов штампа): запрос полей штампа.
      - Неинтерактивный (не-TTY ИЛИ передан хотя бы один флаг штампа):
        штамп заполняется из флагов; незаданные поля — пустые/дефолтные.
    """
    code = args.code.upper()
    name = args.name

    folder_name = f"{code}_{name[:30].replace(' ','_').replace('/','_')}"
    target = PROJECTS_DIR / folder_name

    if target.exists():
        print(warn(f"Папка уже существует: {target}"))
        return

    flag_stamp = {
        "designer":    getattr(args, "designer", None)    or "",
        "checker":     getattr(args, "checker", None)     or "",
        "norm_head":   getattr(args, "norm_head", None)   or "",
        "gip":         getattr(args, "gip", None)         or "",
        "gap":         getattr(args, "gap", None)         or "",
        "org":         getattr(args, "org", None)         or "",
        "city":        getattr(args, "city", None)        or "",
        "stage":       getattr(args, "stage", None)       or "Р",
        "object_type": getattr(args, "object_type", None) or "Общественное здание",
        "system":      getattr(args, "system", None)      or "TN-S",
    }
    has_flags = any(getattr(args, k, None) for k in
                    ("designer","checker","norm_head","gip","gap",
                     "org","city","stage","object_type","system"))
    interactive = sys.stdin.isatty() and not has_flags

    target.mkdir()

    stamp = _input_stamp(flag_stamp) if interactive else flag_stamp
    today = str(__import__("datetime").date.today())

    template = {
        "project": {
            "name":        name,
            "code":        code,
            "object_type": stamp["object_type"],
            "stage":       stamp["stage"],
            "voltage_kv":  0.4,
            "system":      stamp["system"],
            "frequency":   50,
            "city":        stamp["city"],
            "designer":    stamp["designer"],
            "checker":     stamp["checker"],
            "norm_head":   stamp["norm_head"],
            "gip":         stamp["gip"],
            "gap":         stamp["gap"],
            "org":         stamp["org"],
            "revision":    0,
            "date":        today,
            "notes":       "",
            "breaker_series": "IEK",
        },
        "building": {
            "floor_height_m": 3.0,
            "floors":         1,
        },
        "vru": {
            "id":   "ВРУ-1",
            "name": "ВРУ-1",
            "bus_current_a": 250,
            "isc_ka": 10,
            "incoming_cable": {
                "mark":       "ВВГнг-LS",
                "cores":      4,
                "section_mm2":None,
                "length_m":   30,
                "install":    "лоток",
                "ambient_t":  25,
                "parallel":   1,
                "cable_routing": {"mode": "reserve_only", "reserve_pct": 20},
            },
            "feeders": [],
        },
        "outdoor_networks": [],
        "extra_items":      [],
        "changes":          [],
        "_meta": {
            "schema_version": "1.1",
            "created":        today,
            "last_modified":  today,
            "calc_done":      False,
        },
    }

    with open(target / "project.json", "w", encoding="utf-8") as f:
        json.dump(template, f, ensure_ascii=False, indent=2)

    for sub in ["docs", "dwg", "templates"]:
        (target / sub).mkdir()

    print(ok(f"Создан проект: {target}"))
    print(info(f"Редактируй: {target / 'project.json'}"))
    if not interactive and not stamp.get("designer"):
        print(info(f"Заполни штамп: python cli.py stamp {target.name} "
                   f"--field designer --value 'Иванов И.И.'"))


def cmd_validate(args):
    """Проверка корректности project.json."""
    proj_dir = find_project_dir(args.path)
    project = load_project(proj_dir)

    errors = []
    warnings = []

    # ── Обязательные поля проекта ────────────────────────────────────
    for field in ["name", "code", "stage"]:
        if not project.get("project", {}).get(field):
            errors.append(f"project.{field} — не заполнено")

    # ── isc_ka ───────────────────────────────────────────────────────
    isc = project.get("vru", {}).get("isc_ka", 10.0)
    if not (0.1 <= isc <= 50):
        warnings.append(f"vru.isc_ka = {isc} кА — вне диапазона 0.1–50 кА")

    vru = project.get("vru", {})
    if not vru.get("feeders"):
        warnings.append("vru.feeders — пустой список, добавь группы")

    # ── Сбор ID для проверки дублей + проверка значений ─────────────
    all_consumer_ids: list = []
    all_panel_ids: list = []

    for feeder in vru.get("feeders", []):
        for panel in feeder.get("panels", []):
            pid = panel.get("id", "")
            if pid:
                all_panel_ids.append(pid)

            for c in panel.get("consumers", []):
                cid = c.get("id", "")
                if cid:
                    all_consumer_ids.append(cid)

                if not c.get("power_kw"):
                    errors.append(f"  {cid or '?'} — нет мощности power_kw")
                elif c.get("power_kw", 0) <= 0:
                    errors.append(f"  {cid or '?'} — power_kw должна быть > 0")

                if not c.get("cable"):
                    warnings.append(f"  {cid or '?'} — нет данных кабеля")

                if not str(c.get("name", "")).strip():
                    errors.append(f"  {cid or '?'} — пустое наименование")

                cp = c.get("cos_phi")
                if cp is not None and not (0 < float(cp) <= 1.0):
                    errors.append(f"  {cid or '?'} — cos_phi={cp} вне диапазона (0; 1]")

                df = c.get("demand_factor")
                if df is not None and not (0 < float(df) <= 1.0):
                    errors.append(f"  {cid or '?'} — demand_factor={df} вне диапазона (0; 1]")

    # Дубли потребителей
    seen_c: set = set()
    for cid in all_consumer_ids:
        if cid in seen_c:
            errors.append(f"Дублирующийся ID потребителя: {cid}")
        seen_c.add(cid)

    # Дубли щитов
    seen_p: set = set()
    for pid in all_panel_ids:
        if pid in seen_p:
            errors.append(f"Дублирующийся ID щита: {pid}")
        seen_p.add(pid)

    # ── Категорийное соответствие (ПУЭ гл.1.2) ──────────────────────
    try:
        from rules.category_rules import check_all_compliance
        violations = check_all_compliance(project)
        for v in violations:
            errors.append(f"  {v['message']}")
    except Exception:
        pass

    p_name = project.get("project", {}).get("name", "")
    print(hdr(f"Проверка: {p_name}"))

    if errors:
        for e in errors:
            print(err(e))
    if warnings:
        for w in warnings:
            print(warn(w))
    if not errors and not warnings:
        print(ok("Структура корректна"))
    elif not errors:
        print(ok(f"Ошибок нет, предупреждений: {len(warnings)}"))
    else:
        print(err(f"Ошибок: {len(errors)}, предупреждений: {len(warnings)}"))
        sys.exit(1)


def cmd_calc(args):
    """Полный расчёт ВРУ."""
    from calc.engine import calculate_project
    from changes.detector import detect_changes

    proj_dir = find_project_dir(args.path)
    project = load_project(proj_dir)

    p_name = project.get("project", {}).get("name", "")
    print(hdr(f"Расчёт: {p_name}"))

    # Сравниваем с прошлым снапшотом — что изменилось с прошлого расчёта
    old_snapshot = project.get("_vru_snapshot")
    if old_snapshot:
        changed_items = detect_changes({"vru": old_snapshot}, project)
        if changed_items:
            print(f"  {C.CYAN}Изменения с прошлого расчёта:{C.RESET} {len(changed_items)} позиций")
            for it in changed_items[:6]:
                fld = it["field"]
                if fld == "new_consumer":
                    print(f"    + {it['id']} добавлен")
                elif fld == "deleted_consumer":
                    print(f"    − {it['id']} удалён")
                else:
                    print(f"    {it['id']}: {fld}  {it['old']} → {it['new']}")
            if len(changed_items) > 6:
                print(f"    ... ещё {len(changed_items) - 6}")
            print(f"  {C.GRAY}Для регистрации ревизии: python cli.py change {args.path}{C.RESET}")
            print()

    result = calculate_project(project)
    # Обновляем снапшот ВРУ для следующего сравнения
    result["_vru_snapshot"] = copy.deepcopy(result["vru"])
    save_project(result, proj_dir)

    s   = result["_results"]["summary"]
    vru = result["_results"]["vru"]

    print(ok("Расчёт завершён"))
    print()
    print(f"  {'Установленная мощность':<30} {s['p_installed_kw']:>8.2f} кВт")
    print(f"  {'Расчётная мощность':<30} {s['p_calc_kw']:>8.2f} кВт")
    print(f"  {'Расчётная полная мощность':<30} {s['s_calc_kva']:>8.2f} кВА")
    print(f"  {'Расчётный cos φ':<30} {s['cos_phi']:>8.3f}")
    print(f"  {'Расчётный ток ВРУ':<30} {s['i_vru_a']:>8.2f} А")

    cable = s.get("incoming_cable", {})
    if cable.get("section_mm2"):
        print(f"  {'Вводной кабель':<30} {cable['mark']} {cable['cores']}×{cable['section_mm2']} мм²")

    breaker = vru.get("breaker", {})
    if breaker.get("rating"):
        print(f"  {'Вводной автомат':<30} АВ {breaker['rating']}А хар.{breaker['char']}")

    print()
    for feeder in vru.get("feeders", []):
        print(f"  {C.CYAN}{feeder['id']} {feeder['name']}{C.RESET}  "
              f"Pр={feeder['p_calc_kw']:.2f}кВт  Iр={feeder['i_calc_a']:.1f}А")
        for panel in feeder.get("panels", []):
            cb  = panel.get("cable", {})
            br  = panel.get("breaker", {})
            sec = cb.get("section_mm2", "?")
            rat = br.get("rating", "?")
            du  = cb.get("voltage_drop_pct", 0)
            du_warn = f" {C.YELLOW}⚠ΔU={du}%{C.RESET}" if du > 5 else f"  ΔU={du}%"
            print(f"    {panel['id']:<8} {panel['name']:<30}  "
                  f"Iр={panel['i_calc_a']:>6.1f}А  "
                  f"кабель {cb.get('mark','?')} {cb.get('cores','?')}×{sec}мм²  "
                  f"АВ {rat}А{du_warn}")

    # Итог по кабелям
    n_du = n_kzt = n_kzs = 0
    for feeder in vru.get("feeders", []):
        for panel in feeder.get("panels", []):
            for cb_src in [panel.get("cable", {})] + [c.get("cable", {}) for c in panel.get("consumers", [])]:
                if not cb_src.get("du_ok", True):      n_du  += 1
                if not cb_src.get("kz_thermal_ok", True): n_kzt += 1
                if not cb_src.get("kz_sens_ok", True):    n_kzs += 1
    cable_total = n_du + n_kzt + n_kzs
    if cable_total:
        print(warn(f"Кабели: {n_du} ΔU | {n_kzt} термо-КЗ | {n_kzs} чувств. → python cli.py check-cables {args.path}"))
    else:
        print(ok("Кабели: нарушений нет"))

    print()
    print(info(f"Результаты сохранены: {proj_dir / 'project.json'}"))


def cmd_calc_outdoor(args):
    """Расчёт наружных сетей освещения."""
    from outdoor.calc_outdoor import calc_all_outdoor, print_outdoor_report

    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)

    p_name = project.get("project", {}).get("name", "")
    print(hdr(f"Расчёт наружных сетей: {p_name}"))

    if not project.get("outdoor_networks"):
        print(warn("outdoor_networks не заданы в project.json"))
        return

    result = calc_all_outdoor(project)
    save_project(result, proj_dir)

    print_outdoor_report(result)
    print()
    print(info(f"Результаты сохранены: {proj_dir / 'project.json'}"))


def cmd_summary(args):
    """Краткая сводка по рассчитанному проекту."""
    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)

    if not project.get("_results"):
        print(warn("Проект ещё не рассчитан. Запусти: python cli.py calc " + args.path))
        return

    s   = project["_results"]["summary"]
    p   = project["project"]
    vru = project["_results"]["vru"]

    print(hdr(f"{p['code']} — {p['name']}"))
    print(f"  Стадия: {p['stage']}   Ред. {p['revision']}   {p['date']}")
    print(f"  Система: {p['system']}  {p['voltage_kv']} кВ")
    print()
    print(f"  Мощность установленная:  {s['p_installed_kw']:>8.2f} кВт")
    print(f"  Мощность расчётная:      {s['p_calc_kw']:>8.2f} кВт")
    print(f"  Полная мощность:         {s['s_calc_kva']:>8.2f} кВА")
    print(f"  cos φ расчётный:         {s['cos_phi']:>8.3f}")
    print(f"  Ток ВРУ:                 {s['i_vru_a']:>8.2f} А")

    cable = s.get("incoming_cable", {})
    if cable.get("section_mm2"):
        print(f"  Вводной кабель:    {cable['mark']} {cable['cores']}×{cable['section_mm2']} мм²")

    print()
    n_du = n_kzt = n_kzs = 0
    for feeder in vru.get("feeders", []):
        for panel in feeder.get("panels", []):
            if not panel.get("cable", {}).get("ok"):
                print(err(f"  {panel['id']} — {panel['cable'].get('error','ошибка подбора кабеля')}"))
            for cb in ([panel.get("cable", {})]
                       + [c.get("cable", {}) for c in panel.get("consumers", [])]):
                if not cb.get("du_ok", True):         n_du  += 1
                if not cb.get("kz_thermal_ok", True): n_kzt += 1
                if not cb.get("kz_sens_ok", True):    n_kzs += 1

    if n_du or n_kzt or n_kzs:
        print(warn(f"Кабели: {n_du} ΔU | {n_kzt} термо-КЗ | {n_kzs} чувств."
                   f"  →  python cli.py check-cables {args.path}"))
    else:
        print(ok("Кабели: нарушений нет"))

    # Категорийность здания
    bld = project.get("_building", {})
    if bld:
        cat  = bld.get("category_pue", "?")
        desc = bld.get("vru_description", "")
        ok_flag = bld.get("compliance_ok", True)
        viol    = bld.get("violations", 0)
        cat_str = f"Категория здания: {cat} — {desc}"
        if ok_flag:
            print(ok(cat_str))
        else:
            print(warn(f"{cat_str}  ⚠ {viol} нарушений кат. ПУЭ"))


def cmd_number_cables(args):
    """Автонумерация кабелей: КЛ-ЩО1-01, КЛ-ЩС1-02 и т.д."""
    from dwg.number_cables import (update_cable_numbering, print_cable_numbering,
                                   write_cable_numbers_to_dxf)

    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)
    project  = _ensure_calc(project, proj_dir)

    p_name = project.get("project", {}).get("name", "")
    print(hdr(f"Нумерация кабелей: {p_name}"))

    # Нумерация: по DXF (если передан) или по порядку в проекте
    dxf_path = None
    if getattr(args, "dxf", None):
        from pathlib import Path as _P
        dxf_path = _P(args.dxf)
        if not dxf_path.exists():
            print(warn(f"DXF не найден: {dxf_path} — используется порядок из проекта"))
            dxf_path = None

    project = update_cable_numbering(project, dxf_path)
    save_project(project, proj_dir)

    print_cable_numbering(project)

    n = len(project.get("cable_numbering", {}))
    print()
    print(ok(f"Пронумеровано линий: {n}"))

    # Записываем CABLE_NO в DXF-план если он существует
    plan_dxf = proj_dir / "dwg" / "plan_electrical.dxf"
    if plan_dxf.exists():
        try:
            write_cable_numbers_to_dxf(plan_dxf, project["cable_numbering"])
            print(ok(f"Номера записаны в DXF: plan_electrical.dxf"))
        except Exception as e:
            print(warn(f"DXF обновить не удалось: {e}"))

    print()
    print(info("Кабельный журнал теперь покажет номера КЛ. Запусти:"))
    print(f"  python cli.py docs {args.path} --type cable")


def cmd_sld(args):
    """Генерация однолинейной схемы в DXF."""
    from dwg.create_test_sld import create_test_sld

    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)
    project  = _ensure_calc(project, proj_dir)

    p_name = project.get("project", {}).get("name", "")
    print(hdr(f"Однолинейная схема: {p_name}"))

    dwg_dir = proj_dir / "dwg"
    dwg_dir.mkdir(exist_ok=True)

    out = create_test_sld(project, dwg_dir)
    print(ok(f"Однолинейка: {out.name}"))
    print(info(f"Файл: {out}"))


def cmd_update_attribs(args):
    """Синхронизация атрибутов блоков DXF с результатами расчёта."""
    from dwg.update_attribs import update_attribs, add_changes_trapezoid

    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)
    project  = _ensure_calc(project, proj_dir)

    p_name = project.get("project", {}).get("name", "")
    print(hdr(f"Обновление атрибутов DXF: {p_name}"))

    dwg_dir = proj_dir / "dwg"
    updated_total = 0

    for dxf_file in sorted(dwg_dir.glob("*.dxf")):
        n = update_attribs(project, str(dxf_file))
        if n > 0:
            print(ok(f"{dxf_file.name}: обновлено {n} блоков"))
            updated_total += n

            # Трапеция изменений если есть ревизии
            if project.get("project", {}).get("revision", 0) > 0:
                if add_changes_trapezoid(project, str(dxf_file)):
                    print(info(f"  Трапеция изменений добавлена (рев. {project['project']['revision']})"))

    if updated_total == 0:
        print(warn("DXF-файлы не найдены или блоков с ID_TAG нет"))
    else:
        print()
        print(ok(f"Итого обновлено блоков: {updated_total}"))


def cmd_stamp(args):
    """Редактирование штампа проекта с автопересборкой документов."""
    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)
    proj     = project["project"]

    p_name = proj.get("name", "")
    print(hdr(f"Штамп: {p_name}"))

    if args.field and args.value:
        # Одно поле через флаги
        proj[args.field] = args.value
        print(ok(f"{args.field} = {args.value}"))
    else:
        # Интерактивный ввод — текущие значения как дефолт
        stamp = _input_stamp({
            "designer":    proj.get("designer", ""),
            "checker":     proj.get("checker", ""),
            "norm_head":   proj.get("norm_head", ""),
            "gip":         proj.get("gip", ""),
            "gap":         proj.get("gap", ""),
            "org":         proj.get("org", ""),
            "city":        proj.get("city", ""),
            "stage":       proj.get("stage", "Р"),
            "object_type": proj.get("object_type", ""),
            "system":      proj.get("system", "TN-S"),
        })
        for k, v in stamp.items():
            proj[k] = v

    project["project"] = proj
    save_project(project, proj_dir)
    print(ok("Штамп сохранён"))

    # Автоматически пересоздаём документы если они были сгенерированы ранее
    docs_dir = proj_dir / "docs"
    if docs_dir.exists() and any(docs_dir.iterdir()):
        print(info("Пересоздаю документы..."))
        project = _ensure_calc(project, proj_dir)
        fake_args = type("A", (), {"path": str(proj_dir), "type": None, "format": "docx"})()
        cmd_docs(fake_args)


def cmd_docs(args):
    """Генерация документов."""
    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)
    project  = _ensure_calc(project, proj_dir)

    doc_type = getattr(args, "type", None)
    fmt      = getattr(args, "format", "docx") or "docx"
    p_name   = project["project"]["name"]
    print(hdr(f"Генерация документов: {p_name}"))

    docs_dir = proj_dir / "docs"
    docs_dir.mkdir(exist_ok=True)

    def _try_gen(import_path: str, func_name: str, label: str):
        try:
            mod = __import__(import_path, fromlist=[func_name])
            fn  = getattr(mod, func_name)
            path = fn(project, docs_dir)
            print(ok(f"{label}: {path.name}"))
        except ImportError as e:
            print(info(f"{label} — модуль не найден ({e})"))
        except Exception as e:
            print(err(f"{label} — ошибка: {e}"))

    if doc_type in (None, "spec"):
        if fmt in ("docx", "all"):
            _try_gen("docs.gen_spec", "generate_spec",      "Спецификация (docx)")
        if fmt in ("xlsx", "all"):
            _try_gen("docs.gen_spec", "generate_spec_xlsx", "Спецификация (xlsx)")

    if doc_type in (None, "cable"):
        if fmt in ("docx", "all"):
            _try_gen("docs.gen_cable_journal", "generate_cable_journal",
                     "Кабельный журнал (docx)")
        if fmt in ("xlsx", "all"):
            _try_gen("docs.gen_cable_journal", "generate_cable_journal_xlsx",
                     "Кабельный журнал (xlsx)")

    if doc_type in (None, "load"):
        from docs.gen_load_tables import generate_all_load_tables
        files = generate_all_load_tables(project, docs_dir)
        for f in files:
            print(ok(f"Ведомость нагрузок: {f.name}"))
    if doc_type in (None, "work"):
        _try_gen("docs.gen_work_list", "generate_work_list", "Ведомость работ")
    if doc_type in (None, "pnr"):
        _try_gen("docs.gen_pnr",       "generate_pnr",       "Программа ПНР")

    print()
    print(info(f"Документы: {docs_dir}"))


def cmd_plan(args):
    """Генерация DXF-плана."""
    from dwg.gen_plans import generate_section_plan, generate_summary_plan

    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)
    project  = _ensure_calc(project, proj_dir)

    section  = getattr(args, "section", None)
    p_name   = project["project"]["name"]
    print(hdr(f"Генерация плана: {p_name}"))

    dwg_dir  = proj_dir / "dwg"
    dwg_dir.mkdir(exist_ok=True)

    summary = generate_summary_plan(project, dwg_dir)
    print(ok(f"Сводный план: {summary.name}"))

    section_f = generate_section_plan(project, dwg_dir, section)
    label = f"Раздел {section}" if section else "Полный план"
    print(ok(f"{label}: {section_f.name}"))

    print()
    print(info(f"Чертежи: {dwg_dir}"))


def cmd_check_cables(args):
    """Проверка кабелей: ΔU, термостойкость КЗ, чувствительность КЗ."""
    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)
    project  = _ensure_calc(project, proj_dir)

    p_name = project["project"]["name"]
    print(hdr(f"Проверка кабелей: {p_name}"))

    vru = project.get("_results", {}).get("vru", {})
    du_errors   = []
    kz_therm    = []
    kz_sens     = []

    def _collect(obj_id, obj_name, cb, level=""):
        if not cb:
            return
        prefix = f"{level}{obj_id} {obj_name[:30]}"
        du  = cb.get("voltage_drop_pct", 0)
        lim = cb.get("du_limit_pct", 5.0)
        if not cb.get("du_ok", True):
            du_errors.append(f"  {prefix:<44} ΔU={du}% > {lim}%  "
                             f"({cb.get('mark','?')} {cb.get('cores','?')}×{cb.get('section_mm2','?')}мм², "
                             f"L={cb.get('length_m_calc', cb.get('length_m','?'))}м)")
        if not cb.get("kz_thermal_ok", True):
            s_min = cb.get("kz_thermal_s_min_mm2", "?")
            s_act = cb.get("section_mm2", "?")
            kz_therm.append(f"  {prefix:<44} Sмин={s_min}мм² > Sфакт={s_act}мм²  "
                            f"(Iкз={cb.get('isc_ka_source','?')}кА)")
        if not cb.get("kz_sens_ok", True):
            i_end = cb.get("kz_sens_i_end_a", "?")
            i_min = cb.get("kz_sens_i_trip_min_a", "?")
            kz_sens.append(f"  {prefix:<44} Iкз_конец={i_end}А < Iоткл_мин={i_min}А")

    for feeder in vru.get("feeders", []):
        for panel in feeder.get("panels", []):
            _collect(panel["id"], panel["name"], panel.get("cable", {}), "⬦ ")
            for c in panel.get("consumers", []):
                _collect(c["id"], c["name"], c.get("cable", {}), "  · ")

    total = len(du_errors) + len(kz_therm) + len(kz_sens)
    print(f"\nПроверено линий: {sum(1 + len(p.get('consumers',[])) for f in vru.get('feeders',[]) for p in f.get('panels',[]))}")
    print(f"  Нарушений ΔU:              {len(du_errors)}")
    print(f"  Нарушений термостойкости:  {len(kz_therm)}")
    print(f"  Нарушений чувствительности:{len(kz_sens)}")

    if du_errors:
        print(f"\n{C.YELLOW}⚠ Превышение ΔU (обязательно исправить):{C.RESET}")
        for line in du_errors:
            print(line)
    if kz_therm:
        print(f"\n{C.YELLOW}⚠ Термостойкость при КЗ (увеличить сечение):{C.RESET}")
        for line in kz_therm:
            print(line)
    if kz_sens:
        print(f"\n{C.YELLOW}⚠ Чувствительность защиты (проверить автомат):{C.RESET}")
        for line in kz_sens:
            print(line)

    if total == 0:
        print(f"\n{C.GREEN}✓ Нарушений нет{C.RESET}")
    else:
        print()
        print(err(f"Итого нарушений: {total}"))
        sys.exit(1)


def cmd_check_selectivity(args):
    """Проверка селективности автоматов."""
    from rules.selectivity import check_selectivity, print_selectivity_report

    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)
    project  = _ensure_calc(project, proj_dir)

    p_name = project["project"]["name"]
    print(hdr(f"Селективность: {p_name}"))

    results = check_selectivity(project)
    print_selectivity_report(results)

    violations = [r for r in results if not r["ok"]]
    errors_    = [r for r in violations if r["severity"] == "error"]
    if errors_:
        sys.exit(1)


def cmd_check_compensation(args):
    """Проверка необходимости КРМ."""
    from calc.compensation import check_compensation_needed, print_compensation_report, update_compensation

    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)
    project  = _ensure_calc(project, proj_dir)

    p_name = project["project"]["name"]
    print(hdr(f"КРМ: {p_name}"))

    result  = check_compensation_needed(project)
    print_compensation_report(result)

    # Сохраняем результат КРМ в project
    project = update_compensation(project)
    save_project(project, proj_dir)


def cmd_change(args):
    """Регистрация ревизии (изменения) проекта — трапеция ГОСТ 21.1101."""
    from changes.detector import register_change, REASON_CODES

    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)
    p_name   = project.get("project", {}).get("name", "")
    rev_cur  = project.get("project", {}).get("revision", 0)

    print(hdr(f"Регистрация изменения: {p_name}  (текущая ред. {rev_cur})"))

    # Описание
    desc = getattr(args, "desc", None) or ""
    if not desc:
        print("  Коды причин:")
        for code, text in REASON_CODES.items():
            print(f"    {code} — {text}")
        desc = _prompt("Описание изменения", required=True)

    # Код причины
    reason = getattr(args, "reason", None) or ""
    if not reason:
        reason = _prompt("Код причины (01–04)", default="04")
    reason = reason.strip().lstrip("0") or "4"
    reason = reason.zfill(2)
    if reason not in REASON_CODES:
        reason = "04"

    # Категория
    category = getattr(args, "category", None) or "general"

    updated = register_change(
        project,
        description=desc,
        reason_code=reason,
        category=category,
    )
    save_project(updated, proj_dir)

    new_rev = updated["project"]["revision"]
    reason_text = REASON_CODES.get(reason, reason)
    print(ok(f"Ревизия {new_rev} зарегистрирована"))
    print(f"  Описание: {desc}")
    print(f"  Причина:  {reason} — {reason_text}")
    print(f"  Затронуто: {', '.join(updated['changes'][-1]['affected_docs'])}")
    print()
    print(warn("Выполни пересчёт и перегенерируй документы:"))
    print(f"  python cli.py calc {args.path}")
    print(f"  python cli.py docs {args.path}")


def cmd_import(args):
    """Импорт потребителей из DXF или таблицы Excel."""
    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)

    source     = args.source       # "dwg" или "table"
    file_path  = Path(args.file)
    section    = getattr(args, "section", "ТХ")

    p_name = project["project"]["name"]
    print(hdr(f"Импорт ({source}): {p_name}"))

    if not file_path.exists():
        print(err(f"Файл не найден: {file_path}"))
        sys.exit(1)

    # 1. Парсинг
    if source == "dwg":
        from parsers.parse_dwg_assignment import parse_dwg_assignment
        new_consumers = parse_dwg_assignment(file_path, section_code=section)
    elif source == "table":
        from parsers.parse_load_table import parse_load_table
        new_consumers = parse_load_table(file_path, section_code=section)
    else:
        print(err(f"Неизвестный тип источника: {source}. Используй 'dwg' или 'table'"))
        sys.exit(1)

    if not new_consumers:
        print(warn("Потребители не найдены в файле"))
        return

    print(info(f"Найдено {len(new_consumers)} потребителей в разделе {section}"))

    # 2. Показываем что нашли
    for c in new_consumers[:10]:
        reserve_mark = " [резерв]" if c.get("reserve") else ""
        print(f"  {c['id']:<10} {c['name'][:35]:<35} "
              f"P={c.get('power_kw',0):.2f}кВт  кат.{c.get('category_pue',3)}{reserve_mark}")
    if len(new_consumers) > 10:
        print(f"  ... и ещё {len(new_consumers)-10} потребителей")

    # 3. Формируем щиты
    from panels.auto_panels import auto_assign_panels

    # Существующие щиты из первого фидера (или создаём фидер)
    vru = project.setdefault("vru", {})
    feeders = vru.setdefault("feeders", [])

    # Ищем фидер с нужным разделом или создаём
    target_feeder = None
    for f in feeders:
        if f.get("section") == section:
            target_feeder = f
            break
    if target_feeder is None:
        target_feeder = {
            "id":      section,
            "name":    f"Раздел {section}",
            "section": section,
            "panels":  []
        }
        feeders.append(target_feeder)

    existing_panels = target_feeder.get("panels", [])
    updated_panels  = auto_assign_panels(new_consumers, existing_panels)

    # Подсчёт новых
    existing_ids = {c["id"] for p in existing_panels for c in p.get("consumers", [])}
    new_count    = sum(1 for p in updated_panels
                       for c in p.get("consumers", []) if c["id"] not in existing_ids)

    target_feeder["panels"] = updated_panels

    save_project(project, proj_dir)

    print()
    print(ok(f"Добавлено новых потребителей: {new_count}"))
    print(ok(f"Щитов сформировано: {len(updated_panels)}"))
    print(info(f"Запусти расчёт: python cli.py calc {args.path}"))


def cmd_compare_kp(args):
    """Сверка спецификации проекта с КП поставщика."""
    from parsers.compare_kp import compare_kp

    proj_dir = find_project_dir(args.path)
    project  = load_project(proj_dir)

    p_name = project.get("project", {}).get("name", "")
    p_code = project.get("project", {}).get("code", "")
    print(hdr(f"Сверка с КП: {p_name}"))

    # Спецификация требует выполненного расчёта
    if not project.get("_results"):
        print(err("Проект не рассчитан — спецификация недоступна."))
        print(info(f"Запусти: python cli.py calc {args.path}"))
        sys.exit(1)

    kp_file = Path(args.kp_file)
    if not kp_file.exists():
        print(err(f"Файл КП не найден: {kp_file}"))
        sys.exit(1)

    rows = compare_kp(project, str(kp_file))

    if not rows:
        print(warn("Спецификация и КП пусты — нечего сверять."))
        return

    # Раскраска статусов
    _ICON = {"found": "✓", "not_found": "✗", "extra_in_kp": "+"}
    _CLR  = {"found": C.GREEN, "not_found": C.RED, "extra_in_kp": C.YELLOW}

    # Шапка таблицы
    print()
    fmt = "{:<5} {:<28} {:<40} {:>8} {:<6} {:>10} {:>12}  {}"
    print(C.GRAY + fmt.format(
        "Поз.", "Марка", "Наименование", "Кол.", "Ед.", "Цена", "Сумма", "Статус"
    ) + C.RESET)
    print(C.GRAY + "─" * 130 + C.RESET)

    n_found = n_not_found = n_extra = 0
    total = 0.0

    for r in rows:
        status = r["status"]
        if status == "found":
            n_found += 1
        elif status == "not_found":
            n_not_found += 1
        else:
            n_extra += 1
        total += r["amount"]

        icon = _ICON.get(status, "?")
        clr  = _CLR.get(status, C.RESET)
        mark_show = r["mark"] or r["kp_mark"]
        print(fmt.format(
            r["pos"],
            (mark_show or "")[:28],
            (r["name"] or "")[:40],
            f"{r['qty']:.0f}" if float(r['qty']).is_integer() else f"{r['qty']:.2f}",
            (r["unit"] or "")[:6],
            f"{r['price']:.2f}",
            f"{r['amount']:.2f}",
            f"{clr}{icon} {status}{C.RESET}",
        ))

    n_spec = n_found + n_not_found
    print()
    print(f"  Позиций спецификации: {n_spec}  "
          f"({C.GREEN}найдено: {n_found}{C.RESET}, "
          f"{C.RED}не найдено: {n_not_found}{C.RESET})")
    if n_extra:
        print(f"  {C.YELLOW}В КП есть {n_extra} позиций, отсутствующих в спецификации{C.RESET}")
    print(f"  Общая сумма по найденным: {total:,.2f} ₽".replace(",", " "))

    # XLSX-экспорт
    if getattr(args, "xlsx", False):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print(err("openpyxl не установлен — xlsx-экспорт недоступен"))
            return

        docs_dir = proj_dir / "docs"
        docs_dir.mkdir(exist_ok=True)
        out = docs_dir / f"{p_code or 'project'}_kp_compare.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Сверка с КП"

        title_font  = Font(bold=True, size=12)
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        fill_not    = PatternFill("solid", fgColor="FFE0E0")
        fill_extra  = PatternFill("solid", fgColor="FFFFE0")
        thin        = Side(border_style="thin", color="999999")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)
        center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        right       = Alignment(horizontal="right",  vertical="center")

        # Строка 1 — заголовок
        ws.cell(row=1, column=1, value=f"Сверка спецификации с КП — {p_name}").font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        # Строка 2 — шапка таблицы
        headers = ["Поз.", "Марка", "Наименование", "Кол.", "Ед.",
                   "Цена, ₽", "Сумма, ₽", "Статус"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        # Заморозка
        ws.freeze_panes = "A3"

        # Данные
        row_idx = 3
        for r in rows:
            status = r["status"]
            mark_show = r["mark"] or r["kp_mark"]
            row_values = [
                r["pos"], mark_show, r["name"], r["qty"], r["unit"],
                round(r["price"], 2), round(r["amount"], 2),
                {"found": "✓ найдено",
                 "not_found": "✗ не найдено",
                 "extra_in_kp": "+ только в КП"}[status],
            ]
            for col, val in enumerate(row_values, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.border = border
                if col in (1, 4, 5, 6, 7):
                    c.alignment = right if col != 5 else center
                else:
                    c.alignment = left
                if status == "not_found":
                    c.fill = fill_not
                elif status == "extra_in_kp":
                    c.fill = fill_extra
            row_idx += 1

        # Итоговая строка
        ws.cell(row=row_idx, column=1, value="ИТОГО:").font = Font(bold=True)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        c_total = ws.cell(row=row_idx, column=7, value=round(total, 2))
        c_total.font = Font(bold=True)
        c_total.alignment = right
        c_total.border = border
        ws.cell(row=row_idx, column=8,
                value=f"{n_found}/{n_spec}").font = Font(bold=True)
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = border

        # Автоширина колонок
        widths = [6, 28, 50, 10, 8, 14, 16, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 22

        wb.save(out)
        print()
        print(ok(f"XLSX-сверка сохранена: {out}"))


# ── ТОЧКА ВХОДА ───────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Система автоматизации электропроектирования",
        formatter_class=argparse.RawTextHelpFormatter
    )
    sub = parser.add_subparsers(dest="command")

    # list
    sub.add_parser("list", help="Список всех проектов")

    # new
    p_new = sub.add_parser("new", help="Создать новый проект")
    p_new.add_argument("code", help="Код объекта, напр. ОБЪ-2025-001")
    p_new.add_argument("name", help="Название объекта")
    p_new.add_argument("--designer",    help="Разработал (ФИО)")
    p_new.add_argument("--checker",     help="Проверил (ФИО)")
    p_new.add_argument("--norm-head",   dest="norm_head", help="Н.контроль (ФИО)")
    p_new.add_argument("--gip",         help="ГИП (ФИО)")
    p_new.add_argument("--gap",         help="ГАП (ФИО)")
    p_new.add_argument("--org",         help="Организация")
    p_new.add_argument("--city",        help="Город")
    p_new.add_argument("--stage",       help="Стадия [Р]")
    p_new.add_argument("--object-type", dest="object_type", help="Тип объекта")
    p_new.add_argument("--system",      help="Система заземления [TN-S]")

    # calc
    p_calc = sub.add_parser("calc", help="Рассчитать нагрузки ВРУ")
    p_calc.add_argument("path", help="Путь к папке проекта или код")

    # calc-outdoor
    p_co = sub.add_parser("calc-outdoor", help="Рассчитать наружные сети")
    p_co.add_argument("path", help="Путь к папке проекта или код")

    # summary
    p_sum = sub.add_parser("summary", help="Краткая сводка результатов")
    p_sum.add_argument("path", help="Путь к папке проекта или код")

    # validate
    p_val = sub.add_parser("validate", help="Проверить структуру JSON")
    p_val.add_argument("path", help="Путь к папке проекта или код")

    # docs
    p_docs = sub.add_parser("docs", help="Сгенерировать документы")
    p_docs.add_argument("path", help="Путь к папке проекта или код")
    p_docs.add_argument("--type", choices=["spec","cable","load","work","pnr"],
                         help="Тип документа (по умолчанию — все)")
    p_docs.add_argument("--format", choices=["docx","xlsx","all"], default="docx",
                         help="Формат вывода: docx, xlsx или all [docx]")

    # plan
    p_plan = sub.add_parser("plan", help="Сгенерировать DXF-план")
    p_plan.add_argument("path", help="Путь к папке проекта или код")
    p_plan.add_argument("--section", help="Раздел: ОВ, ВК, ТХ, ДУ, ПС, ЭОМ, ЭН")

    # check-cables
    p_cc = sub.add_parser("check-cables", help="Проверить кабели: ΔU, КЗ")
    p_cc.add_argument("path", help="Путь к папке проекта или код")

    # check-selectivity
    p_sel = sub.add_parser("check-selectivity", help="Проверить селективность АВ")
    p_sel.add_argument("path", help="Путь к папке проекта или код")

    # check-compensation
    p_comp = sub.add_parser("check-compensation", help="Проверить необходимость КРМ")
    p_comp.add_argument("path", help="Путь к папке проекта или код")

    # import
    p_imp = sub.add_parser("import", help="Импортировать потребителей из файла")
    p_imp.add_argument("path",    help="Путь к папке проекта или код")
    p_imp.add_argument("source",  choices=["dwg","table"],
                        help="Источник: dwg (DXF-чертёж) или table (Excel/CSV)")
    p_imp.add_argument("file",    help="Путь к файлу")
    p_imp.add_argument("--section", default="ТХ",
                        help="Код раздела (ОВ, ВК, ТХ, ...) [по умолчанию: ТХ]")

    # stamp
    p_stamp = sub.add_parser("stamp", help="Редактировать штамп проекта")
    p_stamp.add_argument("path",    help="Путь к папке проекта или код")
    p_stamp.add_argument("--field", help="Поле для изменения (designer, checker, ...)")
    p_stamp.add_argument("--value", help="Новое значение поля")

    # sld
    p_sld = sub.add_parser("sld", help="Генерировать однолинейную схему (DXF)")
    p_sld.add_argument("path", help="Путь к папке проекта или код")

    # update-attribs
    p_ua = sub.add_parser("update-attribs", help="Синхронизировать атрибуты блоков DXF")
    p_ua.add_argument("path", help="Путь к папке проекта или код")

    # number-cables
    p_nc = sub.add_parser("number-cables", help="Автонумерация кабелей КЛ-ЩО1-01 ...")
    p_nc.add_argument("path", help="Путь к папке проекта или код")
    p_nc.add_argument("--dxf", help="DXF-план для нумерации по координатам (опционально)")

    # change
    p_chg = sub.add_parser("change", help="Зарегистрировать ревизию (изменение) проекта")
    p_chg.add_argument("path",       help="Путь к папке проекта или код")
    p_chg.add_argument("--desc",     help="Описание изменения")
    p_chg.add_argument("--reason",   default="04",
                        help="Код причины: 01 ошибка, 02 экспертиза, 03 задание, 04 иное [04]")
    p_chg.add_argument("--category", default="general",
                        choices=["cable","panel","load","general","calc","drawing"],
                        help="Категория изменения [general]")

    # compare-kp
    p_kp = sub.add_parser("compare-kp",
                          help="Сверка спецификации проекта с КП поставщика")
    p_kp.add_argument("path",    help="Путь к папке проекта или код")
    p_kp.add_argument("kp_file", help="Файл КП поставщика (.xlsx, .xls, .csv)")
    p_kp.add_argument("--xlsx",  action="store_true",
                      help="Сохранить отчёт в docs/{code}_kp_compare.xlsx")

    args = parser.parse_args()

    commands = {
        "list":               cmd_list,
        "new":                cmd_new,
        "calc":               cmd_calc,
        "calc-outdoor":       cmd_calc_outdoor,
        "summary":            cmd_summary,
        "validate":           cmd_validate,
        "docs":               cmd_docs,
        "plan":               cmd_plan,
        "check-cables":       cmd_check_cables,
        "check-selectivity":  cmd_check_selectivity,
        "check-compensation": cmd_check_compensation,
        "import":             cmd_import,
        "stamp":              cmd_stamp,
        "sld":                cmd_sld,
        "update-attribs":     cmd_update_attribs,
        "number-cables":      cmd_number_cables,
        "change":             cmd_change,
        "compare-kp":         cmd_compare_kp,
    }

    if args.command in commands:
        commands[args.command](args)
    else:
        parser.print_help()

if __name__ == "__main__":
    main()
